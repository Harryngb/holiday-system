from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from typing import List
from datetime import datetime
from models import User, LeaveRequest


def generate_report(users: List[User], leave_requests: List[LeaveRequest], current_year: str) -> BytesIO:
    wb = Workbook()
    today_str = datetime.now().strftime("%Y年%m月%d日")

    # ========== Sheet 1: 假期汇总 ==========
    ws = wb.active
    ws.title = "假期汇总"

    title = f"nVision Global Ningbo {current_year}年度假期汇总（{today_str}）"
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="微软雅黑", size=16, bold=True, color="1a3c6e")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    headers = [
        "工号", "姓名", "上年度累计天数", "本年度天数", "总天数",
        "使用天数", "剩余天数", "上年度累计小时", "本年度小时",
        "总小时", "使用小时", "剩余小时",
    ]

    header_fill = PatternFill(start_color="1a3c6e", end_color="1a3c6e", fill_type="solid")
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[2].height = 30

    data_font = Font(name="微软雅黑", size=10)
    for row, u in enumerate(users, 3):
        row_data = [
            u.employee_id, u.name,
            u.last_year_days, u.current_year_days, u.total_leave_days,
            u.used_leave_days, u.remaining_days,
            u.last_year_hours, u.current_year_hours, u.total_leave_hours,
            u.used_leave_hours, u.remaining_hours,
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    col_widths = [12, 12, 16, 12, 10, 12, 12, 16, 12, 10, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # ========== Sheet 2: 假期明细 ==========
    ws2 = wb.create_sheet("假期明细")

    detail_title = f"nVision Global Ningbo {current_year}年度假期明细（{today_str}）"
    ws2.merge_cells("A1:N1")
    dt_cell = ws2["A1"]
    dt_cell.value = detail_title
    dt_cell.font = Font(name="微软雅黑", size=16, bold=True, color="1a3c6e")
    dt_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 40

    detail_headers = [
        "工号", "姓名", "日期范围", "类型", "事由",
        "数量", "单位", "开始时间", "结束时间",
        "状态", "实际数量", "审批人", "审批意见", "审批时间",
    ]

    for col, header in enumerate(detail_headers, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws2.row_dimensions[2].height = 30

    status_map = {"pending": "待审批", "approved": "已通过", "rejected": "已拒绝"}

    def get_actual_quantity(lr):
        if lr.status in ("pending", "rejected"):
            return "-"
        if lr.request_type == "overtime":
            return lr.quantity
        if lr.deduction_type == "不扣假":
            return 0
        if lr.deduction_days is not None:
            return lr.deduction_days
        return lr.quantity

    for row, lr in enumerate(leave_requests, 3):
        date_range = f"{lr.start_date}"
        if lr.end_date and lr.end_date != lr.start_date:
            date_range = f"{lr.start_date} ~ {lr.end_date}"

        req_type = "加班" if lr.request_type == "overtime" else (lr.leave_type or "休假")
        reason = lr.overtime_reason if lr.request_type == "overtime" else lr.reason
        unit_label = "天" if lr.unit == "day" else "小时"
        status_label = status_map.get(lr.status, lr.status)
        actual_qty = get_actual_quantity(lr)
        approver = lr.approver.name if lr.approver else ""
        approval_comment = lr.approval_comment or ""
        updated_at = lr.updated_at.strftime("%Y-%m-%d %H:%M") if lr.updated_at else ""

        row_data = [
            lr.user.employee_id if lr.user else "",
            lr.user.name if lr.user else "",
            date_range, req_type, reason or "",
            lr.quantity, unit_label,
            lr.start_time or "", lr.end_time or "",
            status_label, actual_qty,
            approver, approval_comment, updated_at,
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    detail_col_widths = [12, 12, 24, 10, 14, 8, 8, 10, 10, 10, 10, 12, 20, 18]
    for i, width in enumerate(detail_col_widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
